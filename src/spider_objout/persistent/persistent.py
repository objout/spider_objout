import openpyxl


class Persistent:
    def __init__(self, name='persis.xlsx'):
        """Read or write .xlsx file.

        :param name: the name of the file to be manipulated
        :type name: string
        """
        self.name = name

    def write(self, data=[]):
        """Write data to a .xlsx file.

        :param data: the content to write
        :type data: list
        """
        try:
            self.wb = openpyxl.Workbook()
            self.sheet = self.wb['Sheet']

            for item in data:
                self.sheet.append(list(item))

            self.wb.save(self.name)
        except Exception as e:
            raise e

    def read(self):
        """Read data from a .xlsx file."""
        res = []
        try:
            self.wb = openpyxl.load_workbook(self.name)
            self.sheet = self.wb['Sheet']

            iter = self.sheet.iter_rows(min_col=2)

            for item in iter:
                res.append(item[0].value)
        except Exception as e:
            raise e
        return res
